"""Excel Data Sheet generation (PDR Module 1).

Daloopa-style formatting conventions:
- Hardcoded (reported) numbers: BLUE, each one a hyperlink to its SEC filing
- Calculated lines (margins, FCF): BLACK Excel formulas, live in the workbook
- Negative numbers rendered in parentheses via number formats
- Derived values (e.g. computed Q4) marked with a note styling
"""

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.orm import Session

from finclone.models import Company, FinancialFact, KpiFact
from finclone.pipeline.ingest import current_facts
from finclone.pipeline.normalize import CANONICAL_CONCEPTS

_IS_FLOW = {c.name: c.is_flow for c in CANONICAL_CONCEPTS}

_BLUE = Font(color="0000FF")
_BLUE_ITALIC = Font(color="0000FF", italic=True)  # derived values (computed Q4)
_BOLD = Font(bold=True)
_HEADER = Font(bold=True, size=12)

_FMT_USD = "#,##0;(#,##0)"
_FMT_EPS = "#,##0.00;(#,##0.00)"
_FMT_PCT = "0.0%;(0.0%)"

# (row label, canonical concept, number format) — None concept = blank spacer row
_INCOME_LAYOUT = [
    ("Revenue", "revenue", _FMT_USD),
    ("Cost of Revenue", "cost_of_revenue", _FMT_USD),
    ("Gross Profit", "gross_profit", _FMT_USD),
    ("Research & Development", "research_development", _FMT_USD),
    ("SG&A Expense", "sga_expense", _FMT_USD),
    ("Operating Income", "operating_income", _FMT_USD),
    ("Net Income", "net_income", _FMT_USD),
    ("Diluted EPS", "eps_diluted", _FMT_EPS),
    ("Diluted Shares", "shares_diluted", _FMT_USD),
]
_INCOME_FORMULAS = [  # (label, numerator concept, denominator concept)
    ("Gross Margin %", "gross_profit", "revenue"),
    ("Operating Margin %", "operating_income", "revenue"),
    ("Net Margin %", "net_income", "revenue"),
]
_BALANCE_LAYOUT = [
    ("Cash & Equivalents", "cash_and_equivalents", _FMT_USD),
    ("Total Assets", "total_assets", _FMT_USD),
    ("Total Liabilities", "total_liabilities", _FMT_USD),
    ("Stockholders' Equity", "stockholders_equity", _FMT_USD),
    ("Long-Term Debt", "long_term_debt", _FMT_USD),
]
_CASHFLOW_LAYOUT = [
    ("Operating Cash Flow", "operating_cash_flow", _FMT_USD),
    ("Capital Expenditures", "capex", _FMT_USD),
    ("Stock-Based Compensation", "stock_based_compensation", _FMT_USD),
]


def _load_current_facts(session: Session, company_id: int) -> dict[tuple[str, int, str], FinancialFact]:
    """Latest (post-restatement) fact per (concept, year, period)."""
    rows = session.scalars(select(FinancialFact).where(FinancialFact.company_id == company_id))
    grouped: dict[tuple[str, int, str], list[FinancialFact]] = {}
    for f in rows:
        grouped.setdefault((f.canonical_concept, f.fiscal_year, f.fiscal_period), []).append(f)
    return {key: fact for key, facts in grouped.items() if (fact := current_facts(facts))}


def _columns(facts: dict, period: str, max_years: int = 10,
             year_from: int | None = None, year_to: int | None = None,
             quarters_filter: set[str] | None = None) -> list[tuple[int, str]]:
    """The period columns of the sheet, oldest to newest: (fy, 'FY') for annual,
    (fy, 'Q1'..'Q4') for quarterly. Optional pickers narrow the year range and
    (for quarterly sheets) the quarters included."""
    years = sorted({fy for (_, fy, _) in facts})
    if year_from is not None:
        years = [y for y in years if y >= year_from]
    if year_to is not None:
        years = [y for y in years if y <= year_to]
    years = years[-max_years:]
    if period == "annual":
        return [(fy, "FY") for fy in years]
    wanted = quarters_filter or {"Q1", "Q2", "Q3", "Q4"}
    quarters = [(fy, q) for fy in years for q in ("Q1", "Q2", "Q3", "Q4")
                if q in wanted and any((c, fy, q) in facts for c in _IS_FLOW)]
    return quarters


def _fact_for(facts: dict, concept: str, fy: int, col_period: str) -> FinancialFact | None:
    if col_period == "FY":
        # Annual view: flows have FY durations; balance-sheet items use year-end (Q4)
        period = "FY" if _IS_FLOW.get(concept, True) else "Q4"
        return facts.get((concept, fy, period))
    return facts.get((concept, fy, col_period))


def _write_statement(ws, facts: dict, columns: list[tuple[int, str]],
                     layout: list, formulas: list | None = None) -> dict[str, int]:
    ws.cell(row=1, column=1, value=ws.title).font = _HEADER
    for j, (fy, q) in enumerate(columns, start=2):
        label = f"FY{fy}" if q == "FY" else f"{q} FY{fy}"
        cell = ws.cell(row=2, column=j, value=label)
        cell.font = _BOLD
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(j)].width = 14

    ws.column_dimensions["A"].width = 28
    ws.freeze_panes = "B3"

    concept_rows: dict[str, int] = {}
    row = 3
    for label, concept, fmt in layout:
        ws.cell(row=row, column=1, value=label)
        concept_rows[concept] = row
        for j, (fy, q) in enumerate(columns, start=2):
            fact = _fact_for(facts, concept, fy, q)
            if fact is None:
                continue
            cell = ws.cell(row=row, column=j, value=fact.value)
            cell.number_format = fmt
            cell.font = _BLUE_ITALIC if fact.derived else _BLUE
            cell.hyperlink = fact.source_url
        row += 1

    if formulas:
        row += 1  # spacer
        for label, num, den in formulas:
            if num not in concept_rows or den not in concept_rows:
                continue
            ws.cell(row=row, column=1, value=label)
            for j in range(2, len(columns) + 2):
                col = get_column_letter(j)
                num_ref, den_ref = f"{col}{concept_rows[num]}", f"{col}{concept_rows[den]}"
                cell = ws.cell(row=row, column=j,
                               value=f"=IF({den_ref}=0,\"\",{num_ref}/{den_ref})")
                cell.number_format = _FMT_PCT  # formulas stay default (black) font
            row += 1
    return concept_rows


def _write_cashflow(ws, facts: dict, columns: list[tuple[int, str]],
                    layout: list | None = None) -> None:
    layout = layout if layout is not None else _CASHFLOW_LAYOUT
    concept_rows = _write_statement(ws, facts, columns, layout)
    # Free Cash Flow = OCF - CapEx, as a live black formula (only when both
    # source rows made it into the sheet — the picker may exclude them)
    ocf_row = concept_rows.get("operating_cash_flow")
    capex_row = concept_rows.get("capex")
    if ocf_row is None or capex_row is None:
        return
    row = 3 + len(layout) + 1
    ws.cell(row=row, column=1, value="Free Cash Flow")
    for j in range(2, len(columns) + 2):
        col = get_column_letter(j)
        cell = ws.cell(row=row, column=j, value=f"={col}{ocf_row}-{col}{capex_row}")
        cell.number_format = _FMT_USD


def _write_kpis(ws, session: Session, company_id: int) -> None:
    ws.cell(row=1, column=1, value="Industry KPIs (LLM-extracted, verify against source)").font = _HEADER
    headers = ["KPI", "Period", "Value", "Unit", "Source Quote"]
    widths = [32, 16, 18, 12, 80]
    for j, (h, w) in enumerate(zip(headers, widths), start=1):
        ws.cell(row=2, column=j, value=h).font = _BOLD
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = "A3"

    kpis = session.scalars(
        select(KpiFact).where(KpiFact.company_id == company_id).order_by(KpiFact.name, KpiFact.period)
    )
    for i, k in enumerate(kpis, start=3):
        ws.cell(row=i, column=1, value=k.name)
        ws.cell(row=i, column=2, value=k.period)
        value_cell = ws.cell(row=i, column=3, value=k.value if k.value is not None else k.value_text)
        if k.value is not None:
            value_cell.number_format = _FMT_USD
        value_cell.font = _BLUE
        value_cell.hyperlink = k.source_url
        ws.cell(row=i, column=4, value=k.unit)
        ws.cell(row=i, column=5, value=k.source_quote)


def generate_datasheet(session: Session, company: Company, period: str = "annual",
                       concepts: set[str] | None = None,
                       year_from: int | None = None, year_to: int | None = None,
                       quarters: set[str] | None = None) -> bytes:
    """Build the .xlsx data sheet for a company; returns the file bytes.

    Picker options (PDR Module 1): `concepts` limits line items, `year_from` /
    `year_to` bound the fiscal-year range, `quarters` limits quarterly columns.
    All default to the full sheet.
    """
    facts = _load_current_facts(session, company.id)
    columns = _columns(facts, period, year_from=year_from, year_to=year_to,
                       quarters_filter=quarters)

    def pick(layout: list) -> list:
        return [r for r in layout if concepts is None or r[1] in concepts]

    income = pick(_INCOME_LAYOUT)
    balance = pick(_BALANCE_LAYOUT)
    cashflow = pick(_CASHFLOW_LAYOUT)
    # Ratio formulas only when both of their inputs are on the sheet
    income_formulas = [f for f in _INCOME_FORMULAS
                       if concepts is None or (f[1] in concepts and f[2] in concepts)]

    wb = Workbook()
    ws_income = wb.active
    ws_income.title = "Income Statement"
    _write_statement(ws_income, facts, columns, income, income_formulas)
    if balance:
        _write_statement(wb.create_sheet("Balance Sheet"), facts, columns, balance)
    if cashflow:
        _write_cashflow(wb.create_sheet("Cash Flow"), facts, columns, cashflow)
    _write_kpis(wb.create_sheet("Industry KPIs"), session, company.id)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
