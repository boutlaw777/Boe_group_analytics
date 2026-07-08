"""Custom template validation and Excel generation (PDR Module 4: MCP).

Templates let users define their own row order and custom calculated lines.
Formula expressions are a restricted language: canonical concept names joined
by + and - (e.g. "operating_income + stock_based_compensation"). Formulas
compile to live Excel formulas referencing the template's concept rows.
"""

import re
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from finclone.models import Company
from finclone.pipeline.normalize import CANONICAL_CONCEPTS
from finclone.sheets.datasheet import (
    _BLUE,
    _BLUE_ITALIC,
    _BOLD,
    _FMT_EPS,
    _FMT_USD,
    _HEADER,
    _columns,
    _fact_for,
    _load_current_facts,
    _write_kpis,
)

_VALID_CONCEPTS = {c.name for c in CANONICAL_CONCEPTS}
_TOKEN_RE = re.compile(r"[a-z_]+|[+\-]")


def parse_expression(expression: str) -> list[str]:
    """Tokenize a formula expression; raises ValueError on anything outside
    the restricted grammar (concept names, +, -)."""
    tokens = _TOKEN_RE.findall(expression)
    if "".join(tokens).replace("+", "").replace("-", "") != re.sub(r"[\s+\-]", "", expression):
        raise ValueError(f"Formula contains unsupported characters: {expression!r}")
    if not tokens or tokens[0] in "+-" or tokens[-1] in "+-":
        raise ValueError(f"Malformed formula: {expression!r}")
    for i, tok in enumerate(tokens):
        expected_concept = i % 2 == 0
        if expected_concept and tok in "+-":
            raise ValueError(f"Malformed formula (consecutive operators): {expression!r}")
        if not expected_concept and tok not in "+-":
            raise ValueError(f"Malformed formula (missing operator): {expression!r}")
        if expected_concept and tok not in _VALID_CONCEPTS:
            raise ValueError(
                f"Unknown concept {tok!r} in formula. Valid concepts: {', '.join(sorted(_VALID_CONCEPTS))}"
            )
    return tokens


def validate_config(config: dict) -> None:
    """Raise ValueError if a template config is invalid."""
    rows = config.get("rows")
    if not isinstance(rows, list) or not rows:
        raise ValueError("Template must have a non-empty 'rows' list")

    concepts_present: set[str] = set()
    for row in rows:
        if row.get("type") == "concept":
            concepts_present.add(row.get("concept", ""))

    for i, row in enumerate(rows):
        kind = row.get("type")
        if kind == "spacer":
            continue
        if kind == "concept":
            if row.get("concept") not in _VALID_CONCEPTS:
                raise ValueError(
                    f"Row {i + 1}: unknown concept {row.get('concept')!r}. "
                    f"Valid concepts: {', '.join(sorted(_VALID_CONCEPTS))}"
                )
            if not row.get("label"):
                raise ValueError(f"Row {i + 1}: concept row needs a label")
        elif kind == "formula":
            if not row.get("label"):
                raise ValueError(f"Row {i + 1}: formula row needs a label")
            tokens = parse_expression(row.get("expression", ""))
            for tok in tokens[::2]:
                if tok not in concepts_present:
                    raise ValueError(
                        f"Row {i + 1}: formula references {tok!r}, which is not a "
                        "concept row in this template — add it as a row first"
                    )
        else:
            raise ValueError(f"Row {i + 1}: unknown row type {kind!r}")


def generate_from_template(
    session: Session, company: Company, config: dict, template_name: str,
    period: str = "annual",
) -> bytes:
    """Build a .xlsx using a custom template layout; returns the file bytes."""
    facts = _load_current_facts(session, company.id)
    columns = _columns(facts, period)

    wb = Workbook()
    ws = wb.active
    ws.title = "Custom Model"
    ws.cell(row=1, column=1, value=f"{company.ticker} — {template_name}").font = _HEADER
    for j, (fy, q) in enumerate(columns, start=2):
        label = f"FY{fy}" if q == "FY" else f"{q} FY{fy}"
        cell = ws.cell(row=2, column=j, value=label)
        cell.font = _BOLD
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(j)].width = 14
    ws.column_dimensions["A"].width = 30
    ws.freeze_panes = "B3"

    concept_rows: dict[str, int] = {}
    formula_rows: list[tuple[int, list[str]]] = []  # (row, tokens) — filled second pass

    row = 3
    for entry in config["rows"]:
        if entry["type"] == "spacer":
            row += 1
            continue
        ws.cell(row=row, column=1, value=entry["label"])
        if entry["type"] == "concept":
            concept = entry["concept"]
            concept_rows[concept] = row
            fmt = _FMT_EPS if concept == "eps_diluted" else _FMT_USD
            for j, (fy, q) in enumerate(columns, start=2):
                fact = _fact_for(facts, concept, fy, q)
                if fact is None:
                    continue
                cell = ws.cell(row=row, column=j, value=fact.value)
                cell.number_format = fmt
                cell.font = _BLUE_ITALIC if fact.derived else _BLUE
                cell.hyperlink = fact.source_url
        else:  # formula
            formula_rows.append((row, parse_expression(entry["expression"])))
        row += 1

    # Second pass: formulas can reference concept rows defined above OR below them
    for f_row, tokens in formula_rows:
        for j in range(2, len(columns) + 2):
            col = get_column_letter(j)
            parts: list[str] = []
            for tok in tokens:
                parts.append(tok if tok in "+-" else f"{col}{concept_rows[tok]}")
            cell = ws.cell(row=f_row, column=j, value="=" + "".join(parts))
            cell.number_format = _FMT_USD  # formulas keep the default (black) font

    _write_kpis(wb.create_sheet("Industry KPIs"), session, company.id)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
