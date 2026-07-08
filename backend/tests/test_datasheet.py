from datetime import date
from io import BytesIO

import pytest
from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from finclone.db import Base
from finclone.models import Company, FinancialFact
from finclone.sheets.datasheet import generate_datasheet


@pytest.fixture()
def session():
    engine = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(engine)
    with sessionmaker(bind=engine)() as s:
        yield s


def _fact(company_id: int, concept: str, canonical: str, value: float,
          fy: int, fp: str, derived: bool = False) -> FinancialFact:
    return FinancialFact(
        company_id=company_id, concept=concept, canonical_concept=canonical,
        unit="USD", value=value, fiscal_year=fy, fiscal_period=fp,
        start_date=None, end_date=date(fy, 12, 31), form="10-K",
        accession_number=f"0000-{fy}-{fp}", filed_date=date(fy + 1, 2, 1),
        source_url=f"https://www.sec.gov/example/{fy}", derived=derived,
    )


def test_datasheet_formatting_and_audit_links(session):
    company = Company(cik="0000000001", ticker="TEST", name="Test Co", sector="Software & SaaS")
    session.add(company)
    session.flush()
    session.add_all([
        _fact(company.id, "Revenues", "revenue", 1000.0, 2024, "FY"),
        _fact(company.id, "GrossProfit", "gross_profit", -400.0, 2024, "FY"),
        _fact(company.id, "Assets", "total_assets", 5000.0, 2024, "Q4"),
    ])
    session.commit()

    wb = load_workbook(BytesIO(generate_datasheet(session, company, period="annual")))
    income = wb["Income Statement"]

    assert income["B2"].value == "FY2024"
    revenue_cell = income["B3"]
    assert revenue_cell.value == 1000.0
    assert revenue_cell.font.color.rgb.endswith("0000FF")  # hardcoded = blue
    assert revenue_cell.hyperlink.target == "https://www.sec.gov/example/2024"
    assert "(" in revenue_cell.number_format  # negatives shown in parentheses

    # Margin row is a live black formula, not a hardcoded value
    formula_cells = [c for row in income.iter_rows() for c in row
                     if isinstance(c.value, str) and c.value.startswith("=")]
    assert formula_cells
    # Formula cells keep the default (black/theme) font — never the blue used
    # for hardcoded values. After a save/load round-trip the default font's
    # color is a theme color object, not None, so assert "not blue".
    assert all(
        c.font.color is None or c.font.color.rgb != revenue_cell.font.color.rgb
        for c in formula_cells
    )

    # Balance sheet places year-end (Q4) instants under the annual column.
    # Row 3 = Cash & Equivalents (absent in test data), row 4 = Total Assets.
    balance = wb["Balance Sheet"]
    assert balance["A4"].value == "Total Assets"
    assert balance["B4"].value == 5000.0
    assert balance["B3"].value is None  # no cash fact in the fixture


def test_datasheet_has_all_tabs(session):
    company = Company(cik="0000000002", ticker="EMPT", name="Empty Co", sector=None)
    session.add(company)
    session.commit()
    wb = load_workbook(BytesIO(generate_datasheet(session, company)))
    assert wb.sheetnames == ["Income Statement", "Balance Sheet", "Cash Flow", "Industry KPIs"]
