from datetime import date
from io import BytesIO

import pytest
from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from finclone.db import Base
from finclone.models import Company, FinancialFact
from finclone.sheets.templates import generate_from_template, parse_expression, validate_config


@pytest.fixture()
def session():
    engine = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(engine)
    with sessionmaker(bind=engine)() as s:
        yield s


def test_parse_expression_valid():
    assert parse_expression("operating_income + stock_based_compensation") == [
        "operating_income", "+", "stock_based_compensation",
    ]
    assert parse_expression("operating_cash_flow - capex") == [
        "operating_cash_flow", "-", "capex",
    ]


@pytest.mark.parametrize("bad", [
    "revenue +",                    # trailing operator
    "+ revenue",                    # leading operator
    "revenue + + capex",            # consecutive operators
    "revenue * capex",              # unsupported operator
    "revenue + made_up_concept",    # unknown concept
    "revenue + 500",                # literal numbers not allowed
])
def test_parse_expression_rejects(bad):
    with pytest.raises(ValueError):
        parse_expression(bad)


def test_validate_config_requires_formula_refs_as_rows():
    config = {"rows": [
        {"type": "concept", "label": "Revenue", "concept": "revenue"},
        {"type": "formula", "label": "FCF", "expression": "operating_cash_flow - capex"},
    ]}
    with pytest.raises(ValueError, match="not a concept row"):
        validate_config(config)


def test_validate_config_accepts_good_template():
    validate_config({"rows": [
        {"type": "concept", "label": "Operating Income", "concept": "operating_income"},
        {"type": "concept", "label": "SBC", "concept": "stock_based_compensation"},
        {"type": "spacer"},
        {"type": "formula", "label": "Adjusted EBIT",
         "expression": "operating_income + stock_based_compensation"},
    ]})


def _fact(company_id: int, concept: str, canonical: str, value: float) -> FinancialFact:
    return FinancialFact(
        company_id=company_id, concept=concept, canonical_concept=canonical,
        unit="USD", value=value, fiscal_year=2024, fiscal_period="FY",
        start_date=None, end_date=date(2024, 12, 31), form="10-K",
        accession_number="0000-24", filed_date=date(2025, 2, 1),
        source_url="https://www.sec.gov/example", derived=False,
    )


def test_generate_from_template_custom_rows_and_formula(session):
    company = Company(cik="0000000003", ticker="TMPL", name="Template Co", sector=None)
    session.add(company)
    session.flush()
    session.add_all([
        _fact(company.id, "OperatingIncomeLoss", "operating_income", 900.0),
        _fact(company.id, "ShareBasedCompensation", "stock_based_compensation", 100.0),
    ])
    session.commit()

    config = {"rows": [
        # Custom order: SBC moved to the top (the PDR's example use case)
        {"type": "concept", "label": "Stock-Based Comp", "concept": "stock_based_compensation"},
        {"type": "concept", "label": "Operating Income", "concept": "operating_income"},
        {"type": "formula", "label": "Adjusted EBIT",
         "expression": "operating_income + stock_based_compensation"},
    ]}
    validate_config(config)
    wb = load_workbook(BytesIO(generate_from_template(session, company, config, "My Model")))
    ws = wb["Custom Model"]

    assert ws["A3"].value == "Stock-Based Comp"
    assert ws["B3"].value == 100.0
    assert ws["A4"].value == "Operating Income"
    assert ws["B4"].value == 900.0
    assert ws["A5"].value == "Adjusted EBIT"
    assert ws["B5"].value == "=B4+B3"  # live Excel formula referencing the rows
    assert ws["B3"].hyperlink.target == "https://www.sec.gov/example"
